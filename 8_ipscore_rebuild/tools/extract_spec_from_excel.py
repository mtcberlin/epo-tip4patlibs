#!/usr/bin/env python3
"""Extract the IPScore model spec from the EPO Excel workbook — once.

This script is *not* part of the notebook chain. It runs when the spec needs to be
re-derived, reads `7_ipscore/IPscore_3.01 WORKHORSE.xlsx` and writes
`8_ipscore_rebuild/ipscore_spec.json`. Afterwards module 8 is self-contained: no
notebook ever opens the workbook again.

What it pulls out, and from where:

  'Adapted questions and answers'  (hidden sheet — the machine-readable master)
      A   ID              P_A1 … P_E8, in course order
      E   question text        G   explanation        H   assessment factor
      J:N answer texts (with the "1 - " prefix)   AF:AJ the same, clean
      O   risk flag (1 = this question is a risk driver)
      Q   opportunity flag (-1 = this question is an opportunity driver)
      D   OEK flag (-1 = this answer feeds the cash flow)
      T:X the five OEK parameter values for scores 1…5

  'Financial results'             (the three built-in test patents)
      rows 3-10  their *continuous* OEK values     rows 14-26  financial inputs
      row  28    Excel's own NPV — the acceptance test

The risk / opportunity arithmetic comes from 'RiskOpportunity Calculation':
      opportunity_i = (score − 1) × 0.25     over questions flagged Q = −1
      risk_i        = (5 − score) × −0.25    over questions flagged O =  1
      the profile averages each over its flagged questions only.

Usage:  python3 extract_spec_from_excel.py [--xlsx PATH] [--out PATH] [--check]
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
MODULE = HERE.parent
REPO = MODULE.parent

DEFAULT_XLSX = REPO / "7_ipscore" / "IPscore_3.01 WORKHORSE.xlsx"
DEFAULT_OUT = MODULE / "ipscore_spec.json"

SECTIONS = [
    ("A", "Legal status", "A. Legal status"),
    ("B", "Technology", "B. Technology"),
    ("C", "Market conditions", "C. Market conditions"),
    ("D", "Finance", "D. Finance"),
    ("E", "Strategy", "E. Strategy"),
]

# Our own names for the eight parameters that carry money, plus the symbol each
# one has in the cash-flow chain. The Excel only labels them by question id.
OEK_PARAMS = {
    "B5": ("years_to_market", "T", "Years before the technology can be worked commercially"),
    "C2": ("market_growth", "g", "Annual growth rate of the business area"),
    "C3": ("life_expectancy", "L", "Years the technology stays relevant in the market"),
    "C6": ("extra_turnover_share", "et", "Extra turnover obtainable in the business area"),
    "D1": ("output_maintainable", "D1", "Share of output maintainable *without* the patent"),
    "D2": ("development_cost_share", "D2", "Future development cost, as a share of turnover"),
    "D3": ("production_cost_index", "D3", "Production cost index once the technology is used"),
    "D4": ("investment_index", "D4", "Investment intensity relative to today"),
}

# Columns on 'Adapted questions and answers'
COL_ID, COL_OEK_FLAG, COL_QUESTION = 1, 4, 5
COL_EXPLANATION, COL_FACTOR = 7, 8
COL_ANSWER_FIRST = 10  # J:N
COL_RISK_FLAG, COL_OPPORTUNITY_FLAG = 15, 17  # O, Q
COL_OEK_VALUE_FIRST = 20  # T:X
COL_ANSWER_CLEAN_FIRST = 32  # AF:AJ

TRAILING_VALUE = re.compile(r"\s*\[[^\]]*\]\s*$")  # drop the "… [0,5]" hint


def _clean(text) -> str:
    if text is None:
        return ""
    return TRAILING_VALUE.sub("", " ".join(str(text).split())).strip()


def extract_questions(wb) -> list[dict]:
    ws = wb["Adapted questions and answers"]
    by_id = {}
    for row in range(2, 42):  # 40 questions, one per row
        raw_id = ws.cell(row, COL_ID).value
        if not raw_id or not str(raw_id).startswith("P_"):
            continue
        qid = str(raw_id)[2:]  # "P_A1" -> "A1"
        section = qid[0]
        clean = [_clean(ws.cell(row, COL_ANSWER_CLEAN_FIRST + i).value) for i in range(5)]
        prefixed = [_clean(ws.cell(row, COL_ANSWER_FIRST + i).value) for i in range(5)]
        answers = [c or re.sub(r"^\d\s*-\s*", "", p) for c, p in zip(clean, prefixed)]

        entry = {
            "id": qid,
            "section": section,
            "question": _clean(ws.cell(row, COL_QUESTION).value),
            "explanation": " ".join(str(ws.cell(row, COL_EXPLANATION).value or "").split()),
            "factor": _clean(ws.cell(row, COL_FACTOR).value),
            "answers": answers,
            "is_risk_driver": ws.cell(row, COL_RISK_FLAG).value == 1,
            "is_opportunity_driver": ws.cell(row, COL_OPPORTUNITY_FLAG).value == -1,
            "oek": None,
        }

        if ws.cell(row, COL_OEK_FLAG).value == -1:
            values = [ws.cell(row, COL_OEK_VALUE_FIRST + i).value for i in range(5)]
            if any(v is None for v in values):
                raise ValueError(f"{qid} is flagged as an OEK question but has no value table")
            name, symbol, meaning = OEK_PARAMS[qid]
            entry["oek"] = {
                "param": name,
                "symbol": symbol,
                "meaning": meaning,
                "values": [float(v) for v in values],
            }
        by_id[qid] = entry

    missing = set(OEK_PARAMS) - {q for q, e in by_id.items() if e["oek"]}
    if missing:
        raise ValueError(f"expected OEK questions not flagged in the Excel: {sorted(missing)}")
    return [by_id[q] for q in sorted(by_id, key=lambda q: (q[0], int(q[1:])))]


def extract_test_patents(wb) -> list[dict]:
    """The three built-in patents on 'Financial results', with Excel's own NPV.

    Their OEK values are *continuous* — 2.5 years to market is not one of the five
    discrete answers — so they exercise the cash flow directly, not the score
    lookup. That is exactly what makes them a good acceptance test.
    """
    fr = wb["Financial results"]
    oek_rows = {"B5": 3, "C2": 4, "C3": 5, "C6": 6, "D1": 7, "D2": 8, "D3": 9, "D4": 10}
    patents = []
    for col, name in [("B", "Patent 1"), ("C", "Patent 2"), ("D", "Patent 3")]:
        oek = {OEK_PARAMS[q][0]: float(fr[f"{col}{r}"].value) for q, r in oek_rows.items()}
        financials = {
            "turnover": float(fr[f"{col}14"].value),
            "direct_costs": float(fr[f"{col}15"].value),
            "indirect_costs": float(fr[f"{col}16"].value),
            "depreciation": float(fr[f"{col}17"].value),
            "depreciation_period": float(fr[f"{col}20"].value),
            "business_area_share": float(fr[f"{col}23"].value),
            "discount_rate": float(fr[f"{col}26"].value),
        }
        patents.append(
            {
                "name": name,
                "oek": oek,
                "financials": financials,
                "expected_npv": float(fr[f"{col}28"].value),
            }
        )
    return patents


def build_spec(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    questions = extract_questions(wb)
    counts = {key: sum(1 for q in questions if q["section"] == key) for key, _, _ in SECTIONS}
    return {
        "meta": {
            "model": "EPO IPScore 3.01",
            "source_workbook": "IPscore_3.01 WORKHORSE.xlsx",
            "extracted_by": "8_ipscore_rebuild/tools/extract_spec_from_excel.py",
            "note": (
                "Question texts, answer options and the eight OEK value tables originate "
                "from the EPO IPScore tool. Extracted once so that module 8 runs without "
                "the workbook; re-run the extractor to re-derive."
            ),
            "scoring": {
                "scale": [1, 5],
                "max_points": 5 * len(questions),
                "opportunity_formula": "(score - 1) * 0.25, averaged over opportunity drivers",
                "risk_formula": "(5 - score) * -0.25, averaged over risk drivers",
            },
        },
        "sections": [
            {"key": key, "title": title, "question_count": counts[key]}
            for key, title, _ in SECTIONS
        ],
        "questions": questions,
        "test_patents": extract_test_patents(wb),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--check", action="store_true", help="compare with the committed spec instead of writing")
    args = ap.parse_args()

    spec = build_spec(args.xlsx)
    text = json.dumps(spec, indent=2, ensure_ascii=False) + "\n"

    if args.check:
        current = args.out.read_text(encoding="utf-8")
        same = current == text
        print("spec is up to date" if same else "SPEC DIFFERS from the workbook")
        return 0 if same else 1

    args.out.write_text(text, encoding="utf-8")
    n_oek = sum(1 for q in spec["questions"] if q["oek"])
    n_risk = sum(1 for q in spec["questions"] if q["is_risk_driver"])
    n_opp = sum(1 for q in spec["questions"] if q["is_opportunity_driver"])
    print(f"wrote {args.out.relative_to(REPO)}")
    print(f"  {len(spec['questions'])} questions in {len(spec['sections'])} sections")
    print(f"  {n_oek} carry money (OEK) · {n_risk} risk drivers · {n_opp} opportunity drivers")
    print(f"  {len(spec['test_patents'])} test patents with Excel NPVs")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
