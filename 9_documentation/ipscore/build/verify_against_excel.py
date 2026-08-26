#!/usr/bin/env python3
"""Verify the rendered tools' NPV engine against the EPO IPscore Excel workbook
directly -- the actual algorithm derivation/validation step, as opposed to the
rest of build/ which only carries forward already-authored HTML content.

Two checks:
1. IPscore's calcVAN(): run it (via Node, sandboxed) against the 3 built-in
   test patents on the 'Financial results' sheet of IPscore_3.01 WORKHORSE.xlsx,
   using their *continuous* OEK parameter values straight from the sheet, and
   compare to Excel's own cached NPV (Financial calculations!C28/D28/E28).
2. NPV Target Planner's calcNPV(): since that tool only accepts 1-5 discrete
   scores (not continuous values), it can't be fed the Excel patents directly.
   Instead this cross-checks it against an independent Python re-implementation
   of the identical Excel formula chain, using the same score->value lookup
   tables (read from the 'Adapted questions and answers' sheet, rows P_B5 etc),
   confirming both tools share one correct, Excel-derived engine.

Usage: python3 verify_against_excel.py [--xlsx PATH] [--ipscore-html PATH] [--npv-html PATH]
Exits non-zero if any check fails.
"""
import argparse
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent  # the module folder, 9_documentation/ipscore/
NODE_HARNESS = Path(__file__).resolve().parent / "verify_calcvan.js"


def load_excel_patents(xlsx_path):
    """Read the 3 built-in test patents' OEK+financial inputs and Excel's own
    cached NPV, straight from 'Financial results' (cols B/C/D)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fr = wb["Financial results"]
    patents = []
    for col, name in [("B", "Patent 1"), ("C", "Patent 2"), ("D", "Patent 3")]:
        oek = {
            "B5": fr[f"{col}3"].value, "C2": fr[f"{col}4"].value,
            "C3": fr[f"{col}5"].value, "C6": fr[f"{col}6"].value,
            "D1": fr[f"{col}7"].value, "D2": fr[f"{col}8"].value,
            "D3": fr[f"{col}9"].value, "D4": fr[f"{col}10"].value,
        }
        fin = {
            "turnover": fr[f"{col}14"].value, "directCosts": fr[f"{col}15"].value,
            "indirectCosts": fr[f"{col}16"].value, "depreciation": fr[f"{col}17"].value,
            "deprecPeriod": fr[f"{col}20"].value, "sectorShare": fr[f"{col}23"].value * 100,
            "discountRate": fr[f"{col}26"].value * 100,
        }
        expected_npv = fr[f"{col}28"].value
        patents.append({"name": name, "fin": fin, "oek": oek, "expectedNPV": expected_npv})
    return patents


def load_oek_value_tables(xlsx_path):
    """Read the score(1-5)->continuous-value lookup tables straight from the
    'Adapted questions and answers' sheet (rows labelled P_B5, P_C2, ...)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Adapted questions and answers"]
    tables = {}
    for row in ws.iter_rows(min_row=1, max_row=45):
        label = row[0].value
        if label and str(label).startswith("P_"):
            key = str(label)[2:]  # "P_B5" -> "B5"
            vals = [ws.cell(row=row[0].row, column=c).value for c in range(26, 31)]  # Z:AD
            if all(v is not None for v in vals):
                tables[key] = vals
    return tables


def run_node_harness(html_path, fn_name, cases):
    payload = {"htmlPath": str(html_path), "fnName": fn_name, "cases": cases}
    proc = subprocess.run(
        ["node", str(NODE_HARNESS)], input=json.dumps(payload),
        capture_output=True, text=True, check=True,
    )
    return json.loads(proc.stdout)


def python_reference_npv(fin, oek):
    """Independent re-implementation of the Excel 'Financial calculations'
    formula chain (Revenue/Costs/Investments/Regained/Efficiency/Liquidity/NPV),
    used only to cross-check the NPV Planner's calcNPV() in check 2."""
    BT, DC, IC, DP = fin["turnover"], fin.get("directCosts", 0), fin.get("indirectCosts", 0), fin.get("depreciation", 0)
    depPeriod, S = fin["deprecPeriod"], fin["sectorShare"] / 100
    discount = fin["discountRate"] / 100
    if "costRatio" in fin:  # NPV Planner takes cost/deprec as ratios directly
        shareCosts = fin["costRatio"] / 100
        shareDeprec_pct = fin["deprecPct"]
    else:
        shareCosts = (DC + IC) / BT if BT else 0
        shareDeprec_pct = DP / BT * 100 if BT else 0
    invIndex = depPeriod * shareDeprec_pct / 100
    T, g, L, et = oek["B5"], oek["C2"], oek["C3"], oek["C6"]
    D1, D2, D3, D4 = oek["D1"], oek["D2"], oek["D3"], oek["D4"]

    def frac(y):
        if y <= T:
            return 0
        if y - T < 1:
            return y - T
        if y <= T + L:
            return 1
        if y - T - L < 1:
            return 1 - (y - T - L)
        return 0

    rev = [0] * 11
    for y in range(1, 11):
        rev[y] = frac(y) * et * (1 + g) ** (y - 1) * S * (1 + g) ** y * 100
    y_first = next((y for y in range(1, 11) if rev[y] > 0), 0)
    sum_rev = sum(rev[y] for y in range(1, 11) if y > T and y <= T + 1 + depPeriod)
    avg_rev = sum_rev / depPeriod if depPeriod else 0

    van = 0
    for y in range(1, 11):
        revenue = rev[y]
        fr = frac(y)
        dev_cost = D2 * S * 100 if y <= T else 0
        costs = revenue * D3 * shareCosts + dev_cost
        investments = avg_rev * D4 * invIndex if y == y_first else 0
        inv_reduction = depPeriod * shareDeprec_pct * (1 - D4) * S * (1 + g) ** y if y == y_first else 0
        regained = fr * (1 - shareCosts) * (1 - D1) * S * (1 + g) ** y * 100
        efficiency = fr * shareCosts * (1 - D3) * S * (1 + g) ** y * 100
        liquidity = revenue - costs - investments + regained + efficiency + inv_reduction
        van += liquidity * BT / 100 / (1 + discount) ** y
    return van


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(ROOT / "IPscore_3.01 WORKHORSE.xlsx"))
    ap.add_argument("--ipscore-html", default=str(ROOT / "build" / "dist" / "IPscore_EN_Demo.html"))
    ap.add_argument("--npv-html", default=str(ROOT / "build" / "dist" / "NPV_Target_Planner_EN.html"))
    args = ap.parse_args()

    all_pass = True

    print("=== Check 1: IPscore calcVAN() vs. Excel's 3 built-in test patents ===")
    patents = load_excel_patents(args.xlsx)
    cases = [{"name": p["name"], "fin": p["fin"], "oek": p["oek"], "expectedNPV": p["expectedNPV"]} for p in patents]
    results = run_node_harness(args.ipscore_html, "calcVAN", cases)
    for r in results:
        status = "PASS" if r["pass"] else "FAIL"
        print(f"  {r['name']}: computed={r['computed']:.4f} expected={r['expected']:.4f} diff={r['diff']:.6f} [{status}]")
        all_pass = all_pass and r["pass"]

    print("\n=== Check 2: NPV Planner calcNPV() vs. independent Python reference ===")
    oek_tables = load_oek_value_tables(args.xlsx)
    demo = json.loads((ROOT / "build" / "data" / "npv_demo_en.json").read_text())
    oek_continuous = {k: oek_tables[k][v - 1] for k, v in demo["oekScores"].items()}
    fin = demo["inputs"]
    expected = python_reference_npv(
        {"turnover": fin["bt"], "deprecPeriod": fin["deprecPeriod"], "sectorShare": fin["sectorShare"],
         "discountRate": fin["r"], "costRatio": fin["costRatio"], "deprecPct": fin["deprecPct"]},
        oek_continuous,
    )
    case = {"name": "NPV Planner demo (NovaMed)", "inputs": demo["inputs"], "oekScores": demo["oekScores"], "expectedNPV": expected}
    results2 = run_node_harness(args.npv_html, "calcNPV", [case])
    for r in results2:
        status = "PASS" if r["pass"] else "FAIL"
        print(f"  {r['name']}: computed={r['computed']:.4f} expected={r['expected']:.4f} diff={r['diff']:.6f} [{status}]")
        all_pass = all_pass and r["pass"]

    print("\n" + ("ALL CHECKS PASSED" if all_pass else "SOME CHECKS FAILED"))
    sys.exit(0 if all_pass else 1)


if __name__ == "__main__":
    main()
